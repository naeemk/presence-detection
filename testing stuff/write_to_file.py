from openpyxl import Workbook

def write_data_by_type_to_excel(objects, file_name):
    # Create a new workbook
    wb = Workbook()

    # Select the active worksheet
    ws = wb.active

    # Create a dictionary to store data for each type
    data_by_type = {}

    # Iterate through the objects
    for obj in objects:
        data = obj["data"]
        obj_type = obj["type"]

        # Create a new column if it doesn't exist
        if obj_type not in data_by_type:
            data_by_type[obj_type] = []

        # Append data to the corresponding type
        data_by_type[obj_type].append(data)

    # Write data to separate columns for each type
    for col_index, (obj_type, data) in enumerate(data_by_type.items(), start=1):
        column_letter = chr(65 + col_index)  # Convert index to column letter (A, B, C, ...)
        ws[f'{column_letter}1'] = obj_type  # Write type as column header
        for row_index, value in enumerate(data, start=2):
            ws[f'{column_letter}{row_index}'] = value  # Write data in corresponding column

    # Save the workbook
    wb.save(file_name)

# Sample list of objects
objects = [
    {"data": "Apple", "type": "fruit"},
    {"data": "Banana", "type": "fruit"},
    {"data": "Carrot", "type": "vegetable"},
    {"data": "Broccoli", "type": "vegetable"},
    {"data": "Orange", "type": "fruit"},
    {"data": "Spinach", "type": "vegetable"},
    {"data": "Grapes", "type": "fruit"},
]

# File name to save the workbook
file_name = 'data_by_type.xlsx'

# Call the function
write_data_by_type_to_excel(objects, file_name)
